from encodings import utf_8_sig
import openpyxl
import chardet

# TODO
# Limpar codigo
# Escape das colunas
# Nome padrao da sheet
# Alterar extensao do output

def excel_to_csv(input_path, output_path, skip_rows=1, delimiter=";", **kwargs):

    wb = openpyxl.load_workbook(filename=input_path, data_only=True)
    
    o_sheet = wb.get_sheet_by_name(kwargs["sheet"])
    max_columns = o_sheet.max_column
    max_rows = o_sheet.max_row
    row_list = ""
    
    for row in range(1,max_rows):
        output_row = ""
        for column in range(1,max_columns):
            output_row += str(o_sheet.cell(row=row, column=column).value) + delimiter
            
        output_row = output_row[:-1]
        row_list += output_row + "\n"
        
    
    with open(output_path, "w", encoding="utf-8") as output_file:
        output_file.write(row_list)
                